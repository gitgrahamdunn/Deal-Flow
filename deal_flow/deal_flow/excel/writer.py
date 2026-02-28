from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from deal_flow.excel.formatting import add_top_decile_highlight, style_sheet


def _write_df(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    style_sheet(ws)


def write_workbook(
    out_dir: Path,
    readme_rows: list[dict],
    summary: pd.DataFrame,
    operator_detail: pd.DataFrame,
    production_monthly: pd.DataFrame,
    wells: pd.DataFrame,
    data_quality: pd.DataFrame,
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"
    _write_df(ws_readme, pd.DataFrame(readme_rows))

    for sheet_name, df in [
        ("Summary", summary),
        ("Operator Detail", operator_detail),
        ("Production (Monthly)", production_monthly),
        ("Wells", wells),
        ("Data Quality", data_quality),
    ]:
        ws = wb.create_sheet(sheet_name)
        _write_df(ws, df if not df.empty else pd.DataFrame({"note": ["No rows"]}))
        if sheet_name == "Summary" and not df.empty and "avg_oil_bpd_30d" in df.columns:
            col_idx = list(df.columns).index("avg_oil_bpd_30d") + 1
            from openpyxl.utils import get_column_letter

            add_top_decile_highlight(ws, get_column_letter(col_idx), ws.max_row)

    filename = out_dir / f"deal_flow_{datetime.now().date().isoformat()}.xlsx"
    wb.save(filename)
    return filename
